#!/usr/bin/env python3
"""
Enterprise Linux Hardening Tool - CLI Version
A comprehensive command-line interface for system hardening with blockchain verification
"""

import os
import sys
import sqlite3
import subprocess
import threading
import datetime
import hashlib
import platform
import re
import argparse
import json
from pathlib import Path
import time
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import shutil
import tkinter as tk
from tkinter import filedialog

MODULES = [
    ("Access Control", "access_control.sh"),
    ("Package Management", "package_mgmt.sh"),
    ("Filesystem", "filesystem.sh"),
    ("Services", "services.sh"),
    ("System Maintenance", "system_maintenance.sh"),
    ("Firewall", "firewall.sh"),
    ("Network", "network.sh"),
    ("User Accounts", "user_accounts.sh"),
    ("Logging and Auditing", "logging_auditing.sh"),
]

ROLLBACK_SCRIPTS = {
    "Access Control": "access_control_rollback.sh",
    "Package Management": "package_mgmt_rollback.sh",
    "Filesystem": "filesystem_rollback.sh",
    "Services": "services_rollback.sh",
    "System Maintenance": "system_maintenance_rollback.sh",
    "Firewall": "firewall_rollback.sh",
    "Network": "network_rollback.sh",
    "User Accounts": "user_accounts_rollback.sh",
    "Logging and Auditing": "logging_auditing_rollback.sh",
}

DB_FILE = "/home/kali/hardening.db"
OUTPUT_DIR = "/home/kali/hardening_reports"
SCRIPTS_DIR = "."

@dataclass
class ScanResult:
    policy_id: str
    policy_name: str
    expected_value: str
    current_value: str
    status: str
    module_name: str
    timestamp: str


class Colors:
    """ANSI color codes for terminal output"""
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'


class SimpleBlockchainVerifier:
    def __init__(self, db_conn):
        self.db = db_conn
        self.init_blockchain_table()
    
    def init_blockchain_table(self):
        cursor = self.db.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS blockchain_ledger (
                block_id INTEGER PRIMARY KEY AUTOINCREMENT,
                previous_hash TEXT,
                current_hash TEXT NOT NULL,
                data_hash TEXT NOT NULL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                module_name TEXT,
                action_type TEXT,
                description TEXT
            )
        ''')
        self.db.commit()
    
    def add_to_blockchain(self, data, module_name="", action_type="", description=""):
        try:
            data_hash = hashlib.sha256(str(data).encode()).hexdigest()
            
            cursor = self.db.cursor()
            cursor.execute("SELECT current_hash FROM blockchain_ledger ORDER BY block_id DESC LIMIT 1")
            result = cursor.fetchone()
            previous_hash = result['current_hash'] if result else "0" * 64
            
            combined = previous_hash + data_hash
            current_hash = hashlib.sha256(combined.encode()).hexdigest()
            
            cursor.execute('''
                INSERT INTO blockchain_ledger 
                (previous_hash, current_hash, data_hash, module_name, action_type, description)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (previous_hash, current_hash, data_hash, module_name, action_type, description))
            
            self.db.commit()
            
            return {
                'block_id': cursor.lastrowid,
                'current_hash': current_hash,
                'data_hash': data_hash,
                'timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            print(f"[ERROR] Blockchain error: {e}")
            return None
    
    def verify_chain(self):
        try:
            cursor = self.db.cursor()
            cursor.execute("SELECT * FROM blockchain_ledger ORDER BY block_id")
            blocks = cursor.fetchall()
            
            if not blocks:
                return True, "Chain is empty"
            
            previous_hash = "0" * 64
            
            for block in blocks:
                combined = previous_hash + block['data_hash']
                calculated_hash = hashlib.sha256(combined.encode()).hexdigest()
                
                if calculated_hash != block['current_hash']:
                    return False, f"Chain broken at block {block['block_id']}"
                
                previous_hash = block['current_hash']
            
            return True, f"✓ Blockchain verified ({len(blocks)} blocks intact)"
            
        except Exception as e:
            return False, f"Verification error: {str(e)}"
    
    def get_latest_hash(self):
        cursor = self.db.cursor()
        cursor.execute("SELECT current_hash FROM blockchain_ledger ORDER BY block_id DESC LIMIT 1")
        result = cursor.fetchone()
        return result['current_hash'] if result else None
    
    def view_ledger(self, limit=50):
        cursor = self.db.cursor()
        cursor.execute("""
            SELECT block_id, module_name, action_type, 
                   SUBSTR(current_hash, 1, 16) as short_hash, 
                   timestamp, description
            FROM blockchain_ledger 
            ORDER BY block_id DESC
            LIMIT ?
        """, (limit,))
        return cursor.fetchall()


class HardeningCLI:
    def __init__(self):
        self.conn = None
        self.current_module = None
        self.output_dir = Path(OUTPUT_DIR)
        self.output_dir.mkdir(exist_ok=True)
        
        self.connect_database()
        self.blockchain = SimpleBlockchainVerifier(self.conn)
    
    def connect_database(self):
        """Connect to SQLite database"""
        try:
            os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
            self.conn = sqlite3.connect(DB_FILE, check_same_thread=False)
            self.conn.row_factory = sqlite3.Row
            self.init_database()
            print(f"{Colors.GREEN}[SUCCESS] Database connected: {DB_FILE}{Colors.END}")
            return True
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Cannot connect to database: {e}{Colors.END}")
            return False
    
    def init_database(self):
        """Initialize database tables"""
        cursor = self.conn.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS scan_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                policy_id TEXT NOT NULL,
                policy_name TEXT NOT NULL,
                expected_value TEXT,
                current_value TEXT,
                status TEXT,
                module_name TEXT,
                scan_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS fix_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                policy_id TEXT NOT NULL,
                policy_name TEXT NOT NULL,
                original_value TEXT,
                current_value TEXT,
                status TEXT,
                module_name TEXT,
                fix_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS report_hashes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                hash TEXT NOT NULL,
                module_name TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        self.conn.commit()
        
        # Initialize blockchain if empty
        cursor.execute("SELECT COUNT(*) as count FROM blockchain_ledger")
        if cursor.fetchone()['count'] == 0:
            self.blockchain.add_to_blockchain(
                data="GENESIS_BLOCK",
                module_name="SYSTEM",
                action_type="INIT",
                description="Initial blockchain genesis block"
            )
            print(f"{Colors.CYAN}[INFO] Blockchain initialized{Colors.END}")
    
    def run_command(self, cmd, module_name, action):
        """Run a shell command and capture output"""
        print(f"{Colors.CYAN}{'='*80}{Colors.END}")
        print(f"{Colors.BOLD}Module: {module_name}{Colors.END}")
        print(f"{Colors.BOLD}Action: {action.upper()}{Colors.END}")
        print(f"{Colors.BOLD}Command: {' '.join(cmd)}{Colors.END}")
        print(f"{Colors.CYAN}{'='*80}{Colors.END}")
        
        start_time = time.time()
        
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=300  # 5 minute timeout
            )
            
            elapsed_time = time.time() - start_time
            
            # Print output with colors
            for line in result.stdout.split('\n'):
                if not line.strip():
                    continue
                if '[PASS]' in line:
                    print(f"{Colors.GREEN}{line}{Colors.END}")
                elif '[FAIL]' in line:
                    print(f"{Colors.RED}{line}{Colors.END}")
                elif '[WARN]' in line or '[WARNING]' in line:
                    print(f"{Colors.YELLOW}{line}{Colors.END}")
                elif '[INFO]' in line:
                    print(f"{Colors.CYAN}{line}{Colors.END}")
                elif '[FIXED]' in line:
                    print(f"{Colors.BLUE}{line}{Colors.END}")
                else:
                    print(line)
            
            if result.stderr:
                print(f"{Colors.RED}[STDERR]{Colors.END}")
                print(result.stderr)
            
            print(f"{Colors.CYAN}{'='*80}{Colors.END}")
            print(f"{Colors.BOLD}Exit Code: {result.returncode}{Colors.END}")
            print(f"{Colors.BOLD}Time: {elapsed_time:.2f} seconds{Colors.END}")
            
            # Log to blockchain
            self.blockchain.add_to_blockchain(
                data=f"{module_name}_{action}_{result.returncode}",
                module_name=module_name,
                action_type=action.upper(),
                description=f"Executed {action} on {module_name} with exit code {result.returncode}"
            )
            
            return result.returncode == 0
            
        except subprocess.TimeoutExpired:
            print(f"{Colors.RED}[ERROR] Command timed out after 300 seconds{Colors.END}")
            return False
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Failed to execute command: {e}{Colors.END}")
            return False
    
    def scan_module(self, module_idx):
        """Scan a specific module"""
        if module_idx < 0 or module_idx >= len(MODULES):
            print(f"{Colors.RED}[ERROR] Invalid module index{Colors.END}")
            return False
        
        module_name, script_name = MODULES[module_idx]
        script_path = os.path.join(SCRIPTS_DIR, script_name)
        
        if not os.path.exists(script_path):
            print(f"{Colors.RED}[ERROR] Script not found: {script_path}{Colors.END}")
            return False
        
        self.current_module = module_name
        cmd = ["sudo", "bash", script_path, "scan"]
        
        success = self.run_command(cmd, module_name, "scan")
        
        if success:
            print(f"{Colors.GREEN}[SUCCESS] Scan completed for {module_name}{Colors.END}")
            # Update last scan time in stats
            self.update_stats()
        else:
            print(f"{Colors.RED}[FAILURE] Scan failed for {module_name}{Colors.END}")
        
        return success
    
    def fix_module(self, module_idx):
        """Fix issues in a specific module"""
        if module_idx < 0 or module_idx >= len(MODULES):
            print(f"{Colors.RED}[ERROR] Invalid module index{Colors.END}")
            return False
        
        module_name, script_name = MODULES[module_idx]
        script_path = os.path.join(SCRIPTS_DIR, script_name)
        
        if not os.path.exists(script_path):
            print(f"{Colors.RED}[ERROR] Script not found: {script_path}{Colors.END}")
            return False
        
        # Ask for confirmation
        response = input(f"{Colors.YELLOW}Are you sure you want to fix {module_name}? (yes/no): {Colors.END}").strip().lower()
        if response not in ['yes', 'y']:
            print(f"{Colors.YELLOW}[INFO] Fix operation cancelled{Colors.END}")
            return False
        
        self.current_module = module_name
        cmd = ["sudo", "bash", script_path, "fix"]
        
        success = self.run_command(cmd, module_name, "fix")
        
        if success:
            print(f"{Colors.GREEN}[SUCCESS] Fix completed for {module_name}{Colors.END}")
        else:
            print(f"{Colors.RED}[FAILURE] Fix failed for {module_name}{Colors.END}")
        
        return success
    
    def rollback_module(self, module_idx):
        """Rollback fixes for a specific module"""
        if module_idx < 0 or module_idx >= len(MODULES):
            print(f"{Colors.RED}[ERROR] Invalid module index{Colors.END}")
            return False
        
        module_name = MODULES[module_idx][0]
        rollback_script = ROLLBACK_SCRIPTS.get(module_name)
        
        if not rollback_script or not os.path.exists(rollback_script):
            print(f"{Colors.RED}[ERROR] Rollback script not found for {module_name}{Colors.END}")
            return False
        
        # Ask for confirmation
        response = input(f"{Colors.YELLOW}Are you sure you want to rollback {module_name}? (yes/no): {Colors.END}").strip().lower()
        if response not in ['yes', 'y']:
            print(f"{Colors.YELLOW}[INFO] Rollback operation cancelled{Colors.END}")
            return False
        
        self.current_module = module_name
        cmd = ["sudo", "bash", rollback_script]
        
        success = self.run_command(cmd, module_name, "rollback")
        
        if success:
            print(f"{Colors.GREEN}[SUCCESS] Rollback completed for {module_name}{Colors.END}")
        else:
            print(f"{Colors.RED}[FAILURE] Rollback failed for {module_name}{Colors.END}")
        
        return success
    
    def scan_all(self):
        """Scan all modules"""
        print(f"{Colors.BOLD}Starting scan of all {len(MODULES)} modules...{Colors.END}")
        
        results = []
        for idx in range(len(MODULES)):
            module_name = MODULES[idx][0]
            print(f"\n{Colors.CYAN}[{idx+1}/{len(MODULES)}] Scanning: {module_name}{Colors.END}")
            
            success = self.scan_module(idx)
            results.append((module_name, success))
            
            if idx < len(MODULES) - 1:
                print(f"\n{Colors.YELLOW}Press Enter to continue to next module...{Colors.END}")
                input()
        
        # Print summary
        print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
        print(f"{Colors.BOLD}SCAN SUMMARY{Colors.END}")
        print(f"{Colors.BOLD}{'='*80}{Colors.END}")
        
        success_count = sum(1 for _, success in results if success)
        
        for module_name, success in results:
            status = f"{Colors.GREEN}✓ PASS{Colors.END}" if success else f"{Colors.RED}✗ FAIL{Colors.END}"
            print(f"{module_name:30} {status}")
        
        print(f"\n{Colors.BOLD}Total: {success_count}/{len(MODULES)} modules successful{Colors.END}")
        return success_count == len(MODULES)
    
    def fix_all(self):
        """Fix all modules"""
        print(f"{Colors.BOLD}Starting fix of all {len(MODULES)} modules...{Colors.END}")
        
        # Ask for confirmation
        response = input(f"{Colors.RED}WARNING: This will modify system configuration for ALL modules. Continue? (yes/no): {Colors.END}").strip().lower()
        if response not in ['yes', 'y']:
            print(f"{Colors.YELLOW}[INFO] Fix all operation cancelled{Colors.END}")
            return False
        
        results = []
        for idx in range(len(MODULES)):
            module_name = MODULES[idx][0]
            print(f"\n{Colors.CYAN}[{idx+1}/{len(MODULES)}] Fixing: {module_name}{Colors.END}")
            
            success = self.fix_module(idx)
            results.append((module_name, success))
            
            if idx < len(MODULES) - 1:
                print(f"\n{Colors.YELLOW}Press Enter to continue to next module...{Colors.END}")
                input()
        
        # Print summary
        print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
        print(f"{Colors.BOLD}FIX SUMMARY{Colors.END}")
        print(f"{Colors.BOLD}{'='*80}{Colors.END}")
        
        success_count = sum(1 for _, success in results if success)
        
        for module_name, success in results:
            status = f"{Colors.GREEN}✓ PASS{Colors.END}" if success else f"{Colors.RED}✗ FAIL{Colors.END}"
            print(f"{module_name:30} {status}")
        
        print(f"\n{Colors.BOLD}Total: {success_count}/{len(MODULES)} modules successful{Colors.END}")
        return success_count == len(MODULES)
    
    def rollback_all(self):
        """Rollback all modules"""
        print(f"{Colors.BOLD}Starting rollback of all {len(MODULES)} modules...{Colors.END}")
        
        # Ask for confirmation
        response = input(f"{Colors.RED}WARNING: This will rollback ALL fixes across the system. Continue? (yes/no): {Colors.END}").strip().lower()
        if response not in ['yes', 'y']:
            print(f"{Colors.YELLOW}[INFO] Rollback all operation cancelled{Colors.END}")
            return False
        
        results = []
        for idx in range(len(MODULES)):
            module_name = MODULES[idx][0]
            print(f"\n{Colors.CYAN}[{idx+1}/{len(MODULES)}] Rolling back: {module_name}{Colors.END}")
            
            success = self.rollback_module(idx)
            results.append((module_name, success))
            
            if idx < len(MODULES) - 1:
                print(f"\n{Colors.YELLOW}Press Enter to continue to next module...{Colors.END}")
                input()
        
        # Print summary
        print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
        print(f"{Colors.BOLD}ROLLBACK SUMMARY{Colors.END}")
        print(f"{Colors.BOLD}{'='*80}{Colors.END}")
        
        success_count = sum(1 for _, success in results if success)
        
        for module_name, success in results:
            status = f"{Colors.GREEN}✓ PASS{Colors.END}" if success else f"{Colors.RED}✗ FAIL{Colors.END}"
            print(f"{module_name:30} {status}")
        
        print(f"\n{Colors.BOLD}Total: {success_count}/{len(MODULES)} modules successful{Colors.END}")
        return success_count == len(MODULES)
    
    def show_status(self, module_idx=None):
        """Show scan results status"""
        if not self.conn:
            print(f"{Colors.RED}[ERROR] Database not connected{Colors.END}")
            return
        
        cursor = self.conn.cursor()
        
        if module_idx is not None:
            if module_idx < 0 or module_idx >= len(MODULES):
                print(f"{Colors.RED}[ERROR] Invalid module index{Colors.END}")
                return
            
            module_name = MODULES[module_idx][0]
            
            # Get only latest results for this module
            cursor.execute("""
                SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status 
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    WHERE module_name=?
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                WHERE s1.module_name=?
                ORDER BY s1.policy_id
            """, (module_name, module_name))
            
            rows = cursor.fetchall()
            
            if not rows:
                print(f"{Colors.YELLOW}[INFO] No scan results found for {module_name}{Colors.END}")
                return
            
            print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{Colors.BOLD}SCAN RESULTS: {module_name}{Colors.END}")
            print(f"{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{'ID':<10} {'Policy Name':<40} {'Status':<10} {'Current Value':<20}")
            print(f"{'-'*80}")
            
            for row in rows:
                status_color = Colors.GREEN if row['status'] == 'PASS' else Colors.RED if row['status'] == 'FAIL' else Colors.YELLOW
                print(f"{row['policy_id']:<10} {row['policy_name'][:38]:<40} {status_color}{row['status']:<10}{Colors.END} {str(row['current_value'])[:18]:<20}")
            
            # Calculate stats
            total = len(rows)
            passed = sum(1 for r in rows if r['status'] == 'PASS')
            failed = sum(1 for r in rows if r['status'] == 'FAIL')
            compliance = (passed / (passed + failed) * 100) if (passed + failed) > 0 else 0
            
            print(f"\n{Colors.BOLD}STATISTICS:{Colors.END}")
            print(f"  Total Rules: {total}")
            print(f"  Passed: {passed}")
            print(f"  Failed: {failed}")
            print(f"  Compliance: {compliance:.1f}%")
            
        else:
            # Show overall status for all modules
            cursor.execute("""
                SELECT s1.module_name, 
                       COUNT(*) as total,
                       SUM(CASE WHEN s1.status='PASS' THEN 1 ELSE 0 END) as passed,
                       SUM(CASE WHEN s1.status='FAIL' THEN 1 ELSE 0 END) as failed
                FROM scan_results s1
                INNER JOIN (
                    SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                    FROM scan_results 
                    GROUP BY policy_id, module_name
                ) s2 ON s1.policy_id = s2.policy_id 
                    AND s1.module_name = s2.module_name 
                    AND s1.scan_timestamp = s2.max_timestamp
                GROUP BY s1.module_name
                ORDER BY s1.module_name
            """)
            
            rows = cursor.fetchall()
            
            if not rows:
                print(f"{Colors.YELLOW}[INFO] No scan results found in database{Colors.END}")
                return
            
            print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{Colors.BOLD}OVERALL HARDENING STATUS{Colors.END}")
            print(f"{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{'Module':<25} {'Total':<8} {'Passed':<8} {'Failed':<8} {'Compliance':<12}")
            print(f"{'-'*80}")
            
            total_all = 0
            passed_all = 0
            failed_all = 0
            
            for row in rows:
                total = row['total']
                passed = row['passed'] or 0
                failed = row['failed'] or 0
                compliance = (passed / total * 100) if total > 0 else 0
                
                total_all += total
                passed_all += passed
                failed_all += failed
                
                compliance_bar = '█' * int(compliance / 5) + '░' * (20 - int(compliance / 5))
                print(f"{row['module_name']:<25} {total:<8} {passed:<8} {failed:<8} {compliance:6.1f}% {compliance_bar}")
            
            overall_compliance = (passed_all / total_all * 100) if total_all > 0 else 0
            print(f"\n{Colors.BOLD}OVERALL:{Colors.END}")
            print(f"  Total Rules: {total_all}")
            print(f"  Passed: {passed_all}")
            print(f"  Failed: {failed_all}")
            print(f"  Overall Compliance: {overall_compliance:.1f}%")
    
    def show_fix_history(self, module_idx=None):
        """Show fix history"""
        if not self.conn:
            print(f"{Colors.RED}[ERROR] Database not connected{Colors.END}")
            return
        
        cursor = self.conn.cursor()
        
        if module_idx is not None:
            if module_idx < 0 or module_idx >= len(MODULES):
                print(f"{Colors.RED}[ERROR] Invalid module index{Colors.END}")
                return
            
            module_name = MODULES[module_idx][0]
            
            cursor.execute("""
                SELECT policy_id, policy_name, original_value, current_value, status, fix_timestamp
                FROM fix_history 
                WHERE module_name=?
                ORDER BY fix_timestamp DESC
                LIMIT 50
            """, (module_name,))
            
            rows = cursor.fetchall()
            
            if not rows:
                print(f"{Colors.YELLOW}[INFO] No fix history found for {module_name}{Colors.END}")
                return
            
            print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{Colors.BOLD}FIX HISTORY: {module_name}{Colors.END}")
            print(f"{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{'Policy ID':<12} {'Policy Name':<30} {'Original':<15} {'Current':<15} {'Timestamp':<20}")
            print(f"{'-'*80}")
            
            for row in rows:
                print(f"{row['policy_id']:<12} {row['policy_name'][:28]:<30} {str(row['original_value'])[:13]:<15} {str(row['current_value'])[:13]:<15} {row['fix_timestamp'][:19]:<20}")
            
            print(f"\n{Colors.BOLD}Total fixes: {len(rows)}{Colors.END}")
            
        else:
            cursor.execute("""
                SELECT module_name, COUNT(*) as count, 
                       MAX(fix_timestamp) as last_fix
                FROM fix_history 
                GROUP BY module_name
                ORDER BY module_name
            """)
            
            rows = cursor.fetchall()
            
            if not rows:
                print(f"{Colors.YELLOW}[INFO] No fix history found{Colors.END}")
                return
            
            print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{Colors.BOLD}FIX HISTORY SUMMARY{Colors.END}")
            print(f"{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{'Module':<25} {'Total Fixes':<12} {'Last Fix':<20}")
            print(f"{'-'*80}")
            
            total_fixes = 0
            for row in rows:
                total_fixes += row['count']
                last_fix = row['last_fix'][:19] if row['last_fix'] else "Never"
                print(f"{row['module_name']:<25} {row['count']:<12} {last_fix:<20}")
            
            print(f"\n{Colors.BOLD}Total fixes across all modules: {total_fixes}{Colors.END}")
    
    def generate_pdf_report(self, module_idx=None):
        """Generate comprehensive PDF report matching GUI version"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from reportlab.platypus import PageBreak
            
            if module_idx is not None:
                if module_idx < 0 or module_idx >= len(MODULES):
                    print(f"{Colors.RED}[ERROR] Invalid module index{Colors.END}")
                    return None
                module_name = MODULES[module_idx][0]
            else:
                module_name = None
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            if module_name:
                filename = f"hardening_report_{module_name}_{timestamp}.pdf"
            else:
                filename = f"hardening_report_all_{timestamp}.pdf"
            
            output_path = self.output_dir / filename
            
            print(f"{Colors.CYAN}[INFO] Generating comprehensive PDF report...{Colors.END}")
            print(f"{Colors.CYAN}[INFO] Output: {output_path}{Colors.END}")
            
            # Create PDF document
            doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                                   leftMargin=0.5*inch, rightMargin=0.5*inch,
                                   topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            
            styles = getSampleStyleSheet()
            
            # Custom styles matching GUI version
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                alignment=TA_CENTER,
                spaceAfter=15,
                textColor=colors.HexColor('#1e88e5')
            )
            
            section_style = ParagraphStyle(
                'Section',
                parent=styles['Heading3'],
                fontSize=11,
                spaceBefore=10,
                spaceAfter=6,
                textColor=colors.HexColor('#37474f'),
                fontName='Helvetica-Bold'
            )
            
            normal_small = ParagraphStyle(
                'NormalSmall',
                parent=styles['Normal'],
                fontSize=8,
                leading=10
            )
            
            table_cell_style = ParagraphStyle(
                'TableCell',
                parent=styles['Normal'],
                fontSize=8,
                leading=9,
                wordWrap='CJK'
            )
            
            # Get system information
            def get_system_info():
                try:
                    with open('/etc/os-release', 'r') as f:
                        os_info = {}
                        for line in f:
                            if '=' in line:
                                key, value = line.strip().split('=', 1)
                                os_info[key] = value.strip('"')
                    
                    distribution = os_info.get('NAME', 'Unknown')
                    version = os_info.get('VERSION', 'Unknown')
                    
                    return {
                        'os_name': f"{distribution} {version}",
                        'architecture': platform.machine(),
                        'kernel': platform.release(),
                        'hostname': platform.node(),
                        'distribution_id': os_info.get('ID', 'Unknown').upper()
                    }
                except:
                    return {
                        'os_name': platform.system(),
                        'architecture': platform.machine(),
                        'kernel': platform.release(),
                        'hostname': platform.node(),
                        'distribution_id': platform.system()
                    }
            
            def calculate_compliance_stats(module_name=None):
                cursor = self.conn.cursor()
                
                if module_name:
                    query = """
                        SELECT s1.status, COUNT(*) as count 
                        FROM scan_results s1
                        INNER JOIN (
                            SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                            FROM scan_results 
                            WHERE module_name=?
                            GROUP BY policy_id, module_name
                        ) s2 ON s1.policy_id = s2.policy_id 
                            AND s1.module_name = s2.module_name 
                            AND s1.scan_timestamp = s2.max_timestamp
                        WHERE s1.module_name=?
                        GROUP BY s1.status
                    """
                    cursor.execute(query, (module_name, module_name))
                else:
                    query = """
                        SELECT s1.status, COUNT(*) as count 
                        FROM scan_results s1
                        INNER JOIN (
                            SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                            FROM scan_results 
                            GROUP BY policy_id, module_name
                        ) s2 ON s1.policy_id = s2.policy_id 
                            AND s1.module_name = s2.module_name 
                            AND s1.scan_timestamp = s2.max_timestamp
                        GROUP BY s1.status
                    """
                    cursor.execute(query)
                
                stats = cursor.fetchall()
                
                total = 0
                passed = 0
                failed = 0
                manual = 0
                warning = 0
                
                for stat in stats:
                    count = stat['count']
                    status = stat['status']
                    total += count
                    
                    if status == "PASS":
                        passed += count
                    elif status == "FAIL":
                        failed += count
                    elif status == "MANUAL":
                        manual += count
                    elif status in ["WARN", "WARNING"]:
                        warning += count
                
                if (passed + failed) > 0:
                    compliance_pct = (passed / (passed + failed)) * 100
                else:
                    compliance_pct = 0
                
                if compliance_pct >= 90:
                    risk_level = "LOW"
                elif compliance_pct >= 70:
                    risk_level = "MEDIUM"
                else:
                    risk_level = "HIGH"
                
                return {
                    'total_rules': total,
                    'passed': passed,
                    'failed': failed,
                    'manual': manual,
                    'warnings': warning,
                    'compliance_pct': round(compliance_pct, 1),
                    'risk_level': risk_level
                }
            
            system_info = get_system_info()
            stats = calculate_compliance_stats(module_name)
            
            # Add title
            story.append(Paragraph("Linux Hardening Compliance Report", title_style))
            story.append(Spacer(1, 5))
            
            # Metadata table
            meta_data = [
                ["Report Generated:", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["Operating System:", system_info['os_name']],
                ["Architecture:", system_info['architecture']],
                ["Kernel Version:", system_info['kernel']],
                ["Distribution ID:", system_info['distribution_id']],
                ["Module:", module_name if module_name else 'All Modules'],
                ["Report ID:", f"HARDEN-{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"]
            ]
            
            meta_table_data = []
            for label, value in meta_data:
                meta_table_data.append([
                    Paragraph(f"<b>{label}</b>", table_cell_style),
                    Paragraph(value, table_cell_style)
                ])
            
            meta_table = Table(meta_table_data, colWidths=[1.2*inch, 3.8*inch])
            meta_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
            ]))
            
            story.append(meta_table)
            story.append(Spacer(1, 15))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", section_style))
            
            exec_data = [
                ["Total Rules Scanned:", str(stats['total_rules'])],
                ["Rules Passed:", str(stats['passed'])],
                ["Rules Failed:", str(stats['failed'])],
                ["Warnings:", str(stats['warnings'])],
                ["Overall Compliance:", f"{stats['compliance_pct']}%"],
                ["Risk Level:", f"<font color=\"{'#4caf50' if stats['risk_level'] == 'LOW' else '#ff9800' if stats['risk_level'] == 'MEDIUM' else '#f44336'}\">{stats['risk_level']}</font>"]
            ]
            
            exec_table_data = []
            for label, value in exec_data:
                exec_table_data.append([
                    Paragraph(f"<b>{label}</b>", table_cell_style),
                    Paragraph(value, table_cell_style)
                ])
            
            exec_table = Table(exec_table_data, colWidths=[1.5*inch, 1*inch])
            exec_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ]))
            
            story.append(exec_table)
            story.append(Spacer(1, 15))
            
            # Detailed Results
            story.append(Paragraph("Detailed Results", section_style))
            story.append(Spacer(1, 5))
            
            cursor = self.conn.cursor()
            
            if module_name:
                cursor.execute("""
                    SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status 
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        WHERE module_name=?
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    WHERE s1.module_name=?
                    ORDER BY s1.policy_id
                """, (module_name, module_name))
            else:
                cursor.execute("""
                    SELECT s1.module_name, s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status 
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    ORDER BY s1.module_name, s1.policy_id
                """)
            
            rows = cursor.fetchall()
            
            if not rows:
                story.append(Paragraph("No scan results available.", normal_small))
            else:
                if module_name:
                    table_data = []
                    
                    header_row = [
                        Paragraph("<b>ID</b>", table_cell_style),
                        Paragraph("<b>Policy Name</b>", table_cell_style),
                        Paragraph("<b>Expected Value</b>", table_cell_style),
                        Paragraph("<b>Current Value</b>", table_cell_style),
                        Paragraph("<b>Status</b>", table_cell_style)
                    ]
                    table_data.append(header_row)
                    
                    for row in rows:
                        policy_name = str(row['policy_name'])
                        if len(policy_name) > 60:
                            policy_name = policy_name[:57] + "..."
                        
                        expected = str(row['expected_value'] or "")
                        if len(expected) > 30:
                            expected = expected[:27] + "..."
                        
                        current = str(row['current_value'] or "")
                        if len(current) > 30:
                            current = current[:27] + "..."
                        
                        table_data.append([
                            Paragraph(str(row['policy_id']), table_cell_style),
                            Paragraph(policy_name, table_cell_style),
                            Paragraph(expected, table_cell_style),
                            Paragraph(current, table_cell_style),
                            Paragraph(str(row['status']), table_cell_style)
                        ])
                    
                    col_widths = [0.4*inch, 2.5*inch, 1.2*inch, 1.2*inch, 0.6*inch]
                    
                else:
                    table_data = []
                    
                    header_row = [
                        Paragraph("<b>Module</b>", table_cell_style),
                        Paragraph("<b>ID</b>", table_cell_style),
                        Paragraph("<b>Policy Name</b>", table_cell_style),
                        Paragraph("<b>Status</b>", table_cell_style),
                        Paragraph("<b>Expected</b>", table_cell_style),
                        Paragraph("<b>Current</b>", table_cell_style)
                    ]
                    table_data.append(header_row)
                    
                    for row in rows:
                        module = str(row['module_name'])
                        policy_name = str(row['policy_name'])
                        if len(policy_name) > 40:
                            policy_name = policy_name[:37] + "..."
                        
                        expected = str(row['expected_value'] or "")
                        if len(expected) > 15:
                            expected = expected[:12] + "..."
                        
                        current = str(row['current_value'] or "")
                        if len(current) > 15:
                            current = current[:12] + "..."
                        
                        table_data.append([
                            Paragraph(module, table_cell_style),
                            Paragraph(str(row['policy_id']), table_cell_style),
                            Paragraph(policy_name, table_cell_style),
                            Paragraph(str(row['status']), table_cell_style),
                            Paragraph(expected, table_cell_style),
                            Paragraph(current, table_cell_style)
                        ])
                    
                    col_widths = [0.7*inch, 0.4*inch, 2.2*inch, 0.6*inch, 0.8*inch, 0.8*inch]
                
                table = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e88e5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('TOPPADDING', (0, 0), (-1, 0), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
                    ('TOPPADDING', (0, 1), (-1, -1), 3),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ])
                
                status_col_index = 4 if module_name else 3
                
                for i in range(1, len(table_data)):
                    status_cell = table_data[i][status_col_index]
                    status_text = str(status_cell)
                    
                    if "PASS" in status_text:
                        bg_color = colors.HexColor('#4caf50')
                    elif "FAIL" in status_text:
                        bg_color = colors.HexColor('#f44336')
                    elif "MANUAL" in status_text:
                        bg_color = colors.HexColor('#ff9800')
                    elif "WARN" in status_text or "WARNING" in status_text:
                        bg_color = colors.HexColor('#ffc107')
                    else:
                        bg_color = colors.HexColor('#9e9e9e')
                    
                    style.add('BACKGROUND', (status_col_index, i), (status_col_index, i), bg_color)
                    style.add('TEXTCOLOR', (status_col_index, i), (status_col_index, i), colors.white)
                
                table.setStyle(style)
                story.append(table)
                
                if len(rows) > 20:
                    story.append(PageBreak())
                
                story.append(Spacer(1, 10))
                
                # Compliance Summary
                story.append(Paragraph("Compliance Summary", section_style))
                story.append(Spacer(1, 5))
                
                if module_name:
                    summary_data = [
                        ["Status", "Count", "Percentage"],
                        ["Passed", str(stats['passed']), f"{stats['compliance_pct']}%"],
                        ["Failed", str(stats['failed']), f"{100 - stats['compliance_pct']}%"],
                        ["Warnings", str(stats['warnings']), f"{(stats['warnings']/stats['total_rules']*100):.1f}%" if stats['total_rules'] > 0 else "0%"],
                        ["Total", str(stats['total_rules']), "100%"]
                    ]
                else:
                    cursor.execute("""
                        SELECT s1.module_name, 
                               COUNT(*) as total,
                               SUM(CASE WHEN s1.status='PASS' THEN 1 ELSE 0 END) as passed,
                               SUM(CASE WHEN s1.status='FAIL' THEN 1 ELSE 0 END) as failed
                        FROM scan_results s1
                        INNER JOIN (
                            SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                            FROM scan_results 
                            GROUP BY policy_id, module_name
                        ) s2 ON s1.policy_id = s2.policy_id 
                            AND s1.module_name = s2.module_name 
                            AND s1.scan_timestamp = s2.max_timestamp
                        GROUP BY s1.module_name
                        ORDER BY s1.module_name
                    """)
                    module_stats = cursor.fetchall()
                    
                    summary_data = [["Module", "Total", "Passed", "Failed", "Compliance"]]
                    for stat in module_stats:
                        total = stat['total']
                        passed = stat['passed'] or 0
                        failed = stat['failed'] or 0
                        compliance = (passed / total * 100) if total > 0 else 0
                        
                        summary_data.append([
                            stat['module_name'],
                            str(total),
                            str(passed),
                            str(failed),
                            f"{compliance:.1f}%"
                        ])
                    
                    summary_data.append([
                        "TOTAL",
                        str(stats['total_rules']),
                        str(stats['passed']),
                        str(stats['failed']),
                        f"{stats['compliance_pct']}%"
                    ])
                
                summary_table_data = []
                for i, row in enumerate(summary_data):
                    row_data = []
                    for j, cell in enumerate(row):
                        if i == 0:
                            row_data.append(Paragraph(f"<b>{cell}</b>", table_cell_style))
                        else:
                            row_data.append(Paragraph(str(cell), table_cell_style))
                    summary_table_data.append(row_data)
                
                summary_table = Table(summary_table_data, colWidths=[1.5*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.8*inch])
                summary_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#607d8b')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                ])
                
                if not module_name:
                    summary_style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e3f2fd'))
                    summary_style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
                
                summary_table.setStyle(summary_style)
                story.append(summary_table)
            
            # Integrity Verification
            story.append(Spacer(1, 15))
            story.append(Paragraph("="*80, normal_small))
            
            report_data = str(rows) + str(datetime.datetime.now()) + str(stats)
            report_hash = hashlib.sha256(report_data.encode()).hexdigest()
            
            integrity_text = f"""
            <b>Integrity Verification:</b><br/>
            <font size="7">Document Hash: {report_hash[:32]}...<br/>
            Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
            System: {system_info['hostname']}<br/>
            To verify: Compare this hash with stored hash in database.</font>
            """
            story.append(Paragraph(integrity_text, normal_small))
            
            doc.build(story)
            
            # Store report hash in database
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS report_hashes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    filename TEXT NOT NULL,
                    hash TEXT NOT NULL,
                    module_name TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            cursor.execute('''
                INSERT INTO report_hashes (filename, hash, module_name) 
                VALUES (?, ?, ?)
            ''', (str(output_path), report_hash, module_name))
            
            self.conn.commit()
            
            # Add to blockchain
            self.blockchain.add_to_blockchain(
                data=report_hash,
                module_name=module_name or "ALL",
                action_type="PDF_REPORT",
                description=f"Comprehensive PDF report generated: {filename}"
            )
            
            print(f"{Colors.GREEN}[SUCCESS] Comprehensive PDF report generated: {output_path}{Colors.END}")
            print(f"{Colors.GREEN}[INFO] Report ID: HARDEN-{timestamp}{Colors.END}")
            print(f"{Colors.GREEN}[INFO] Integrity Hash: {report_hash[:16]}...{Colors.END}")
            
            return str(output_path)
            
        except ImportError as e:
            print(f"{Colors.RED}[ERROR] Required PDF libraries not installed: {e}{Colors.END}")
            print(f"{Colors.YELLOW}[INFO] Install with: pip install reportlab{Colors.END}")
            return None
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Failed to generate PDF: {e}{Colors.END}")
            return None
    
    def verify_pdf_report(self):
        """Verify the integrity of a PDF report"""
        # Create a hidden root window for file dialog
        root = tk.Tk()
        root.withdraw()
        
        filename = filedialog.askopenfilename(
            title="Select PDF Report to Verify",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        root.destroy()
        
        if not filename:
            print(f"{Colors.YELLOW}[INFO] No file selected{Colors.END}")
            return
        
        try:
            with open(filename, 'rb') as f:
                content = f.read()
            
            content_str = content.decode('latin-1', errors='ignore')
            
            hash_match = re.search(r'Hash: ([a-fA-F0-9]{64})', content_str)
            
            if not hash_match:
                print(f"{Colors.RED}[ERROR] No hash found in PDF{Colors.END}")
                return
            
            file_hash = hash_match.group(1)
            
            cursor = self.conn.cursor()
            cursor.execute('''
                SELECT hash FROM report_hashes 
                WHERE filename=? OR hash LIKE ?
                ORDER BY created_at DESC LIMIT 1
            ''', (filename, f"{file_hash[:20]}%"))
            
            result = cursor.fetchone()
            
            if result:
                db_hash = result['hash']
                if file_hash == db_hash:
                    print(f"{Colors.GREEN}✓ Report is authentic (not tampered){Colors.END}")
                else:
                    print(f"{Colors.RED}✗ Report has been modified!{Colors.END}")
            else:
                print(f"{Colors.RED}✗ Report not found in database{Colors.END}")
                
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Verification error: {str(e)}{Colors.END}")
    
    def export_to_excel(self, module_idx=None):
        """Export results to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            if module_idx is not None:
                if module_idx < 0 or module_idx >= len(MODULES):
                    print(f"{Colors.RED}[ERROR] Invalid module index{Colors.END}")
                    return None
                module_name = MODULES[module_idx][0]
            else:
                module_name = None
            
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            if module_name:
                filename = f"hardening_{module_name}_{timestamp}.xlsx"
            else:
                filename = f"hardening_all_{timestamp}.xlsx"
            
            output_path = self.output_dir / filename
            
            print(f"{Colors.CYAN}[INFO] Exporting to Excel...{Colors.END}")
            print(f"{Colors.CYAN}[INFO] Output: {output_path}{Colors.END}")
            
            wb = Workbook()
            ws = wb.active
            
            if module_name:
                ws.title = module_name[:31]
                headers = ["Policy ID", "Policy Name", "Expected Value", "Current Value", "Status", "Timestamp"]
            else:
                ws.title = "All Results"
                headers = ["Module", "Policy ID", "Policy Name", "Expected Value", "Current Value", "Status", "Timestamp"]
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
            
            cursor = self.conn.cursor()
            
            if module_name:
                cursor.execute("""
                    SELECT s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status, s1.scan_timestamp
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        WHERE module_name=?
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    WHERE s1.module_name=?
                    ORDER BY s1.policy_id
                """, (module_name, module_name))
            else:
                cursor.execute("""
                    SELECT s1.module_name, s1.policy_id, s1.policy_name, s1.expected_value, s1.current_value, s1.status, s1.scan_timestamp
                    FROM scan_results s1
                    INNER JOIN (
                        SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                        FROM scan_results 
                        GROUP BY policy_id, module_name
                    ) s2 ON s1.policy_id = s2.policy_id 
                        AND s1.module_name = s2.module_name 
                        AND s1.scan_timestamp = s2.max_timestamp
                    ORDER BY s1.module_name, s1.policy_id
                """)
            
            rows = cursor.fetchall()
            
            # Write data
            for row_num, row in enumerate(rows, 2):
                if module_name:
                    row_data = (row['policy_id'], row['policy_name'], row['expected_value'], 
                               row['current_value'], row['status'], row['scan_timestamp'])
                else:
                    row_data = (row['module_name'], row['policy_id'], row['policy_name'], 
                               row['expected_value'], row['current_value'], row['status'], 
                               row['scan_timestamp'])
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    
                    # Color code status column
                    if (module_name and col_num == 5) or (not module_name and col_num == 6):
                        if value == "PASS":
                            cell.fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
                        elif value == "FAIL":
                            cell.fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
                        elif value == "MANUAL":
                            cell.fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            
            # Add to blockchain
            self.blockchain.add_to_blockchain(
                data=f"excel_export_{filename}",
                module_name=module_name or "ALL",
                action_type="EXCEL_EXPORT",
                description=f"Excel report exported: {filename}"
            )
            
            print(f"{Colors.GREEN}[SUCCESS] Excel file exported: {output_path}{Colors.END}")
            print(f"{Colors.GREEN}[INFO] Total records exported: {len(rows)}{Colors.END}")
            return str(output_path)
            
        except ImportError as e:
            print(f"{Colors.RED}[ERROR] Required Excel libraries not installed: {e}{Colors.END}")
            print(f"{Colors.YELLOW}[INFO] Install with: pip install openpyxl{Colors.END}")
            return None
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Failed to export to Excel: {e}{Colors.END}")
            return None
    
    def verify_blockchain(self):
        """Verify blockchain integrity"""
        if not self.conn:
            print(f"{Colors.RED}[ERROR] Database not connected{Colors.END}")
            return
        
        is_valid, message = self.blockchain.verify_chain()
        
        if is_valid:
            print(f"{Colors.GREEN}{message}{Colors.END}")
        else:
            print(f"{Colors.RED}{message}{Colors.END}")
        
        # Show blockchain stats
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) as count FROM blockchain_ledger")
        count = cursor.fetchone()['count']
        print(f"\n{Colors.BOLD}Blockchain Statistics:{Colors.END}")
        print(f"  Total blocks: {count}")
        
        if count > 0:
            cursor.execute("SELECT MIN(timestamp) as first, MAX(timestamp) as last FROM blockchain_ledger")
            times = cursor.fetchone()
            print(f"  First block: {times['first']}")
            print(f"  Last block: {times['last']}")
        
        return is_valid
    
    def view_blockchain(self, limit=20):
        """View blockchain ledger"""
        if not self.conn:
            print(f"{Colors.RED}[ERROR] Database not connected{Colors.END}")
            return
        
        blocks = self.blockchain.view_ledger(limit)
        
        if not blocks:
            print(f"{Colors.YELLOW}[INFO] No blockchain transactions found{Colors.END}")
            return
        
        print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
        print(f"{Colors.BOLD}BLOCKCHAIN LEDGER (Last {limit} transactions){Colors.END}")
        print(f"{Colors.BOLD}{'='*80}{Colors.END}")
        print(f"{'Block':<8} {'Module':<20} {'Action':<15} {'Hash':<20} {'Timestamp':<20}")
        print(f"{'-'*80}")
        
        for block in blocks:
            print(f"{block['block_id']:<8} {block['module_name'][:18]:<20} {block['action_type'][:13]:<15} "
                  f"{block['short_hash']}... {block['timestamp'][:19]:<20}")
    
    def cleanup_database(self):
        """Clean up duplicate records"""
        if not self.conn:
            print(f"{Colors.RED}[ERROR] Database not connected{Colors.END}")
            return
        
        response = input(f"{Colors.YELLOW}This will remove duplicate scan results, keeping only the latest. Continue? (yes/no): {Colors.END}").strip().lower()
        if response not in ['yes', 'y']:
            print(f"{Colors.YELLOW}[INFO] Cleanup cancelled{Colors.END}")
            return
        
        try:
            print(f"{Colors.CYAN}[INFO] Cleaning up database...{Colors.END}")
            
            cursor = self.conn.cursor()
            
            # Count before cleanup
            cursor.execute("SELECT COUNT(*) as count FROM scan_results")
            before_count = cursor.fetchone()['count']
            
            # Create temporary table with unique records
            cursor.execute('''
                CREATE TEMPORARY TABLE temp_scan_results AS
                SELECT 
                    MIN(id) as id,
                    policy_id,
                    policy_name,
                    expected_value,
                    current_value,
                    status,
                    module_name,
                    MAX(scan_timestamp) as scan_timestamp
                FROM scan_results
                GROUP BY policy_id, module_name
            ''')
            
            # Delete all records
            cursor.execute('DELETE FROM scan_results')
            
            # Insert unique records back
            cursor.execute('''
                INSERT INTO scan_results 
                (id, policy_id, policy_name, expected_value, current_value, status, module_name, scan_timestamp)
                SELECT id, policy_id, policy_name, expected_value, current_value, status, module_name, scan_timestamp
                FROM temp_scan_results
            ''')
            
            # Drop temporary table
            cursor.execute('DROP TABLE temp_scan_results')
            
            # Count after cleanup
            cursor.execute("SELECT COUNT(*) as count FROM scan_results")
            after_count = cursor.fetchone()['count']
            
            self.conn.commit()
            
            removed = before_count - after_count
            print(f"{Colors.GREEN}[SUCCESS] Database cleanup completed!{Colors.END}")
            print(f"  Records before: {before_count}")
            print(f"  Records after: {after_count}")
            print(f"  Duplicates removed: {removed}")
            
            # Log to blockchain
            self.blockchain.add_to_blockchain(
                data=f"db_cleanup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}",
                module_name="SYSTEM",
                action_type="DB_CLEANUP",
                description=f"Database cleanup: removed {removed} duplicate records"
            )
            
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Failed to clean up database: {e}{Colors.END}")
    
    def update_stats(self):
        """Update statistics"""
        if not self.conn:
            return
        
        cursor = self.conn.cursor()
        
        # Get total unique policies
        cursor.execute("""
            SELECT COUNT(DISTINCT s1.policy_id || s1.module_name) as unique_policies
            FROM scan_results s1
            INNER JOIN (
                SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                FROM scan_results 
                GROUP BY policy_id, module_name
            ) s2 ON s1.policy_id = s2.policy_id 
                AND s1.module_name = s2.module_name 
                AND s1.scan_timestamp = s2.max_timestamp
        """)
        total_policies = cursor.fetchone()[0] or 0
        
        # Get passed policies
        cursor.execute("""
            SELECT COUNT(*) as passed
            FROM scan_results s1
            INNER JOIN (
                SELECT policy_id, module_name, MAX(scan_timestamp) as max_timestamp
                FROM scan_results 
                GROUP BY policy_id, module_name
            ) s2 ON s1.policy_id = s2.policy_id 
                AND s1.module_name = s2.module_name 
                AND s1.scan_timestamp = s2.max_timestamp
            WHERE s1.status='PASS'
        """)
        passed = cursor.fetchone()['passed'] or 0
        
        compliance = (passed / total_policies * 100) if total_policies > 0 else 0
        
        print(f"\n{Colors.BOLD}Current Statistics:{Colors.END}")
        print(f"  Total policies: {total_policies}")
        print(f"  Passed: {passed}")
        print(f"  Failed: {total_policies - passed}")
        print(f"  Compliance: {compliance:.1f}%")
    
    def show_system_info(self):
        """Display system information"""
        print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
        print(f"{Colors.BOLD}SYSTEM INFORMATION{Colors.END}")
        print(f"{Colors.BOLD}{'='*80}{Colors.END}")
        
        try:
            with open('/etc/os-release', 'r') as f:
                os_info = {}
                for line in f:
                    if '=' in line:
                        key, value = line.strip().split('=', 1)
                        os_info[key] = value.strip('"')
            
            print(f"{'System':<20}: {platform.node()}")
            print(f"{'OS':<20}: {os_info.get('NAME', 'Unknown')} {os_info.get('VERSION', '')}")
            print(f"{'Kernel':<20}: {platform.release()}")
            print(f"{'Architecture':<20}: {platform.machine()}")
            print(f"{'Python':<20}: {platform.python_version()}")
            print(f"{'Database':<20}: {DB_FILE}")
            print(f"{'Scripts Directory':<20}: {SCRIPTS_DIR}")
            print(f"{'Output Directory':<20}: {OUTPUT_DIR}")
            
        except Exception as e:
            print(f"{Colors.RED}[ERROR] Failed to get system info: {e}{Colors.END}")
    
    def interactive_menu(self):
        """Display interactive menu"""
        while True:
            print(f"\n{Colors.BOLD}{'='*80}{Colors.END}")
            print(f"{Colors.BOLD}ENTERPRISE LINUX HARDENING TOOL - CLI VERSION{Colors.END}")
            print(f"{Colors.BOLD}{'='*80}{Colors.END}")
            
            print(f"\n{Colors.BOLD}Modules:{Colors.END}")
            for idx, (name, script) in enumerate(MODULES):
                print(f"  {idx+1:2}. {name}")
            
            print(f"\n{Colors.BOLD}Actions:{Colors.END}")
            print(f"  1.  Scan a module")
            print(f"  2.  Fix a module")
            print(f"  3.  Rollback a module")
            print(f"  4.  Scan all modules")
            print(f"  5.  Fix all modules")
            print(f"  6.  Rollback all modules")
            print(f"  7.  Show status")
            print(f"  8.  Show fix history")
            print(f"  9.  Generate PDF report")
            print(f"  10. Export to Excel")
            print(f"  11. Verify blockchain")
            print(f"  12. View blockchain")
            print(f"  13. Cleanup database")
            print(f"  14. Show system info")
            print(f"  15. Update statistics")
            print(f"  16. Verify PDF report")
            print(f"  17. Exit")
            
            try:
                choice = input(f"\n{Colors.BOLD}Enter your choice (1-17): {Colors.END}").strip()
                
                if choice == '17' or choice.lower() in ['exit', 'quit']:
                    print(f"\n{Colors.CYAN}Exiting... Goodbye!{Colors.END}")
                    break
                
                elif choice == '1':  # Scan a module
                    idx = self.get_module_index()
                    if idx is not None:
                        self.scan_module(idx)
                
                elif choice == '2':  # Fix a module
                    idx = self.get_module_index()
                    if idx is not None:
                        self.fix_module(idx)
                
                elif choice == '3':  # Rollback a module
                    idx = self.get_module_index()
                    if idx is not None:
                        self.rollback_module(idx)
                
                elif choice == '4':  # Scan all modules
                    self.scan_all()
                
                elif choice == '5':  # Fix all modules
                    self.fix_all()
                
                elif choice == '6':  # Rollback all modules
                    self.rollback_all()
                
                elif choice == '7':  # Show status
                    idx = self.get_module_index(allow_none=True)
                    self.show_status(idx)
                
                elif choice == '8':  # Show fix history
                    idx = self.get_module_index(allow_none=True)
                    self.show_fix_history(idx)
                
                elif choice == '9':  # Generate PDF report
                    idx = self.get_module_index(allow_none=True)
                    self.generate_pdf_report(idx)
                
                elif choice == '10':  # Export to Excel
                    idx = self.get_module_index(allow_none=True)
                    self.export_to_excel(idx)
                
                elif choice == '11':  # Verify blockchain
                    self.verify_blockchain()
                
                elif choice == '12':  # View blockchain
                    try:
                        limit = input(f"{Colors.BOLD}Number of transactions to show (default 20): {Colors.END}").strip()
                        limit = int(limit) if limit.isdigit() else 20
                        self.view_blockchain(limit)
                    except ValueError:
                        self.view_blockchain()
                
                elif choice == '13':  # Cleanup database
                    self.cleanup_database()
                
                elif choice == '14':  # Show system info
                    self.show_system_info()
                
                elif choice == '15':  # Update statistics
                    self.update_stats()
                
                elif choice == '16':  # Verify PDF report
                    self.verify_pdf_report()
                
                else:
                    print(f"{Colors.RED}[ERROR] Invalid choice: {choice}{Colors.END}")
                    
            except KeyboardInterrupt:
                print(f"\n\n{Colors.YELLOW}Interrupted by user{Colors.END}")
                break
            except Exception as e:
                print(f"{Colors.RED}[ERROR] An error occurred: {e}{Colors.END}")
    
    def get_module_index(self, allow_none=False):
        """Get module index from user input"""
        while True:
            try:
                if allow_none:
                    prompt = f"{Colors.BOLD}Enter module number (1-{len(MODULES)}) or press Enter for all: {Colors.END}"
                else:
                    prompt = f"{Colors.BOLD}Enter module number (1-{len(MODULES)}): {Colors.END}"
                
                input_str = input(prompt).strip()
                
                if allow_none and input_str == '':
                    return None
                
                idx = int(input_str) - 1
                if 0 <= idx < len(MODULES):
                    return idx
                else:
                    print(f"{Colors.RED}[ERROR] Invalid module number. Must be between 1 and {len(MODULES)}{Colors.END}")
                    
            except ValueError:
                print(f"{Colors.RED}[ERROR] Please enter a valid number{Colors.END}")
            except KeyboardInterrupt:
                return None


def main():
    """Main entry point"""
    # Check if running as root for operations that need it
    if os.geteuid() != 0:
        print(f"{Colors.RED}[ERROR] This tool must be run as root for scanning/fixing operations!{Colors.END}")
        print(f"{Colors.YELLOW}[INFO] Please run with: sudo python3 hardening_cli.py{Colors.END}")
        sys.exit(1)
    
    # Create argument parser for command-line mode
    parser = argparse.ArgumentParser(description='Enterprise Linux Hardening Tool - CLI Version')
    parser.add_argument('--scan', type=int, help='Scan specific module (1-9)')
    parser.add_argument('--fix', type=int, help='Fix specific module (1-9)')
    parser.add_argument('--rollback', type=int, help='Rollback specific module (1-9)')
    parser.add_argument('--scan-all', action='store_true', help='Scan all modules')
    parser.add_argument('--fix-all', action='store_true', help='Fix all modules')
    parser.add_argument('--rollback-all', action='store_true', help='Rollback all modules')
    parser.add_argument('--status', type=int, nargs='?', const=None, help='Show status (module number or empty for all)')
    parser.add_argument('--history', type=int, nargs='?', const=None, help='Show fix history (module number or empty for all)')
    parser.add_argument('--pdf', type=int, nargs='?', const=None, help='Generate PDF report (module number or empty for all)')
    parser.add_argument('--excel', type=int, nargs='?', const=None, help='Export to Excel (module number or empty for all)')
    parser.add_argument('--verify-blockchain', action='store_true', help='Verify blockchain integrity')
    parser.add_argument('--view-blockchain', type=int, nargs='?', const=20, help='View blockchain ledger (number of transactions)')
    parser.add_argument('--cleanup', action='store_true', help='Cleanup database duplicates')
    parser.add_argument('--system-info', action='store_true', help='Show system information')
    parser.add_argument('--stats', action='store_true', help='Show current statistics')
    parser.add_argument('--verify-pdf', action='store_true', help='Verify PDF report integrity')
    parser.add_argument('--interactive', '-i', action='store_true', help='Launch interactive menu')
    
    args = parser.parse_args()
    
    # Create CLI instance
    cli = HardeningCLI()
    
    # Check if any arguments were provided
    if any(vars(args).values()):
        # Command-line mode
        if args.scan:
            cli.scan_module(args.scan - 1)
        elif args.fix:
            cli.fix_module(args.fix - 1)
        elif args.rollback:
            cli.rollback_module(args.rollback - 1)
        elif args.scan_all:
            cli.scan_all()
        elif args.fix_all:
            cli.fix_all()
        elif args.rollback_all:
            cli.rollback_all()
        elif args.status is not None:
            cli.show_status(args.status - 1 if args.status != 0 else None)
        elif args.history is not None:
            cli.show_fix_history(args.history - 1 if args.history != 0 else None)
        elif args.pdf is not None:
            cli.generate_pdf_report(args.pdf - 1 if args.pdf != 0 else None)
        elif args.excel is not None:
            cli.export_to_excel(args.excel - 1 if args.excel != 0 else None)
        elif args.verify_blockchain:
            cli.verify_blockchain()
        elif args.view_blockchain:
            cli.view_blockchain(args.view_blockchain)
        elif args.cleanup:
            cli.cleanup_database()
        elif args.system_info:
            cli.show_system_info()
        elif args.stats:
            cli.update_stats()
        elif args.verify_pdf:
            cli.verify_pdf_report()
        elif args.interactive:
            cli.interactive_menu()
    else:
        # No arguments, launch interactive mode
        cli.interactive_menu()


if __name__ == "__main__":
    main()
